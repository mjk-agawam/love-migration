"""Convert Government and ERG tabs from raw_source.xlsx to Phase 2 CSV format.

Reads the master Google Sheets export and writes source_p2_government_*.csv and
source_p2_erg_*.csv files that sync.py Steps 7-9 can consume.

Phase 2 CSV columns (24):
  External ID, Name, Location, Area of interest, Account Information,
  Documentation and Requirements, Grant Amount, Due Date, EIN,
  Application Status, Outreach, Research_Links__c, Internal_Notes__c,
  Relationship, Contact 1 NameRaw, Contact 2 NameRaw, Contact 3 NameRaw,
  Title1 NameRaw, Title2 NameRaw, Title3 NameRaw, Email1 NameRaw,
  Email2 NameRaw, Email3 NameRaw, Source_Sheet__c

Usage:
    python3 convert_gov_erg.py
"""
import csv
import os
import re
import unicodedata

import openpyxl

PROJ = os.path.dirname(os.path.abspath(__file__))
RAW  = os.path.join(PROJ, 'raw_source.xlsx')

P2_HEADERS = [
    'External ID', 'Name', 'Location', 'Area of interest',
    'Account Information', 'Documentation and Requirements',
    'Grant Amount', 'Due Date', 'EIN', 'Application Status',
    'Outreach', 'Research_Links__c', 'Internal_Notes__c',
    'Relationship',
    'Contact 1 NameRaw', 'Contact 2 NameRaw', 'Contact 3 NameRaw',
    'Title1 NameRaw', 'Title2 NameRaw', 'Title3 NameRaw',
    'Email1 NameRaw', 'Email2 NameRaw', 'Email3 NameRaw',
    'Source_Sheet__c',
]


def slugify(text):
    """Generate URL-safe slug from text."""
    text = str(text).strip().lower()
    text = unicodedata.normalize('NFKD', text)
    text = re.sub(r'[^\w\s-]', '', text)
    text = re.sub(r'[\s_]+', '', text)
    return text or 'unknown'


def cell(val):
    """Clean a cell value to a stripped string."""
    if val is None:
        return ''
    return str(val).strip()


def clean_multiline(val):
    """Collapse newlines in a cell to semicolons for CSV safety."""
    s = cell(val)
    if not s:
        return ''
    return re.sub(r'\s*[\r\n]+\s*', '; ', s).strip('; ')


def extract_linkedin(text):
    """Extract a LinkedIn profile URL from text, return (cleaned_text, url)."""
    if not text:
        return text, ''
    match = re.search(r'(?:https?://)?(?:www\.)?linkedin\.com/in/[a-zA-Z0-9_-]+/?', text, re.IGNORECASE)
    if match:
        url = match.group(0)
        if not url.startswith('http'):
            url = 'https://' + url
        cleaned = text[:match.start()] + text[match.end():]
        cleaned = re.sub(r'\s*\|\s*', ' ', cleaned).strip()
        return cleaned, url
    return text, ''


def first_email(text):
    """Extract the first email address from a messy string."""
    if not text:
        return ''
    match = re.search(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', text)
    return match.group(0) if match else ''


def split_contacts(name_str, title_str, email_str, sep=';'):
    """Split multi-value contact fields into list of (name, title, email) tuples.

    Returns up to 3 contacts.
    """
    names = [s.strip() for s in re.split(r'[;\n]+', cell(name_str)) if s.strip()]
    titles = [s.strip() for s in re.split(r'[;\n]+', cell(title_str)) if s.strip()]
    emails = [s.strip() for s in re.split(r'[;\n]+', cell(email_str)) if s.strip()]

    contacts = []
    for i in range(max(len(names), 1)):
        n = names[i] if i < len(names) else ''
        t = titles[i] if i < len(titles) else ''
        e = emails[i] if i < len(emails) else ''
        if n or e:
            contacts.append((n, t, first_email(e) if e else ''))
    return contacts[:3]


def make_row(ext_id, name, source_sheet, **kwargs):
    """Build a Phase 2 row dict."""
    row = {h: '' for h in P2_HEADERS}
    row['External ID'] = ext_id
    row['Name'] = name
    row['Source_Sheet__c'] = source_sheet
    row['Relationship'] = kwargs.get('relationship', 'Government')
    row['Location'] = kwargs.get('location', '')
    row['Area of interest'] = kwargs.get('area', '')
    row['Account Information'] = kwargs.get('acct_info', '')
    row['Documentation and Requirements'] = kwargs.get('docs_reqs', '')
    row['Grant Amount'] = kwargs.get('grant_amount', '')
    row['Due Date'] = kwargs.get('due_date', '')
    row['EIN'] = kwargs.get('ein', '')
    row['Application Status'] = kwargs.get('app_status', '')
    row['Outreach'] = kwargs.get('outreach', '')
    row['Research_Links__c'] = kwargs.get('research_links', '')
    row['Internal_Notes__c'] = kwargs.get('notes', '')

    contacts = kwargs.get('contacts', [])
    for i, (cn, ct, ce) in enumerate(contacts[:3]):
        idx = i + 1
        row[f'Contact {idx} NameRaw' if idx != 2 else 'Contact 2 NameRaw'] = cn
        row[f'Title{idx} NameRaw'] = ct
        row[f'Email{idx} NameRaw'] = ce
    return row


def write_p2_csv(filename, rows):
    """Write rows to a Phase 2 CSV file."""
    path = os.path.join(PROJ, filename)
    with open(path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=P2_HEADERS)
        writer.writeheader()
        writer.writerows(rows)
    print(f'  Wrote {len(rows)} rows → {filename}')
    return path


# ── Government Tab Converters ─────────────────────────────────────────────────

def convert_council_members(wb):
    """Merge Council Members TOTAL (contacts) + Outreach (notes) tabs.

    Account = one per council district, named "NYC Council District {N} - {Member Name}".
    Contacts: Council Member + up to 2 staff (Chief of Staff, Legislative Director).
    """
    ws_total = wb['Gov-Council Members - TOTAL ']
    ws_outreach = wb['Gov-Council Members - Outreach ']

    # Build outreach lookup by district
    outreach_map = {}  # district → {outreach, status, notes, etc.}
    for row in ws_outreach.iter_rows(min_row=2, values_only=True):
        if not any(c for c in row if c is not None and str(c).strip()):
            continue
        district = cell(row[0])  # "District 1"
        outreach_map[district] = {
            'outreach': clean_multiline(row[8]),   # Outreach
            'status': clean_multiline(row[5]),      # Funding Status
            'notes': clean_multiline(row[11]),      # Email Response Status
            'action': clean_multiline(row[12]),     # Action Items
            'meeting': clean_multiline(row[7]),     # Meeting Conducted
            'program': clean_multiline(row[6]),     # L.O.V.E. Program
        }

    rows = []
    for row in ws_total.iter_rows(min_row=2, values_only=True):
        if not any(c for c in row if c is not None and str(c).strip()):
            continue
        borough = cell(row[0])
        district = cell(row[1])  # "District 1"
        school = cell(row[2])
        status = cell(row[3])
        member = cell(row[4])
        member_email = first_email(cell(row[5]))
        address = clean_multiline(row[6])
        phone = cell(row[7])
        cos = cell(row[8])         # Chief of Staff
        cos_email = first_email(cell(row[9]))
        leg_dir = cell(row[10])    # Legislative Director
        leg_email = first_email(cell(row[11]))

        if not member:
            continue

        # District number for slug
        dist_num = re.sub(r'[^0-9]', '', district) or district
        name = f"NYC Council {district} - {member}"
        ext_id = slugify(name)

        # Merge outreach info
        out_info = outreach_map.get(district, {})

        contacts = []
        if member:
            contacts.append((member, 'Council Member', member_email))
        if cos:
            contacts.append((cos, 'Chief of Staff', cos_email))
        if leg_dir:
            contacts.append((leg_dir, 'Legislative Director', leg_email))

        # Build account info from school + status + funding
        acct_parts = []
        if school:
            acct_parts.append(f"L.O.V.E. School: {school}")
        if status:
            acct_parts.append(f"Status: {status}")
        if out_info.get('program'):
            acct_parts.append(f"Program: {out_info['program']}")
        if address:
            acct_parts.append(f"District Office: {address}")
        if phone:
            acct_parts.append(f"Phone: {phone}")
        acct_info = '; '.join(acct_parts)

        # Notes from outreach tab
        notes_parts = []
        if out_info.get('status'):
            notes_parts.append(f"Funding: {out_info['status']}")
        if out_info.get('meeting'):
            notes_parts.append(f"Meeting: {out_info['meeting']}")
        if out_info.get('notes'):
            notes_parts.append(out_info['notes'])
        if out_info.get('action'):
            notes_parts.append(f"Action: {out_info['action']}")

        rows.append(make_row(
            ext_id=ext_id,
            name=name,
            source_sheet='Government Council Members',
            location=f"{borough}, New York City",
            acct_info=acct_info,
            app_status=out_info.get('status', status),
            outreach=out_info.get('outreach', ''),
            notes='; '.join(notes_parts),
            contacts=contacts,
        ))

    return write_p2_csv('source_p2_government_council.csv', rows)


def convert_congress_members(wb):
    """Convert Congress Members tab.

    Data is misaligned: member name in col 3 (Status), email in col 4/5.
    """
    ws = wb['Gov-Congress Members']
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(c for c in row if c is not None and str(c).strip()):
            continue
        borough = cell(row[0])
        district = cell(row[1])
        school = cell(row[2])

        # Member name is in col 3 (misaligned from "Status" header)
        member = cell(row[3])
        if not member:
            continue

        # Col 4 sometimes has email, col 5 has website
        col4 = cell(row[4])
        website = cell(row[5])  # Actually the website/email
        if not website and '@' in col4:
            website = ''

        # Col 6 = District Office Address (actually a contact name)
        contact1_name = cell(row[6])
        contact1_email = first_email(cell(row[7]))  # District Office Phone = actually email

        cos = cell(row[8])
        cos_email = first_email(cell(row[9]))

        leg_dir = cell(row[10])
        leg_email = first_email(cell(row[11]))

        name = f"US Congress District {district} - {member}"
        ext_id = slugify(name)

        contacts = []
        if contact1_name:
            contacts.append((contact1_name, 'District Staff', contact1_email))
        if cos:
            contacts.append((cos, 'Chief of Staff', cos_email))
        if leg_dir:
            # Clean up name
            leg_clean = re.sub(r'\(.*?\)', '', leg_dir).strip()
            contacts.append((leg_clean, 'Legislative Director', leg_email))

        acct_parts = []
        if school:
            acct_parts.append(f"L.O.V.E. School: {school}")
        if website:
            acct_parts.append(f"Website: {website}")

        rows.append(make_row(
            ext_id=ext_id,
            name=name,
            source_sheet='Government Congress Members',
            location=f"{borough}, New York",
            acct_info='; '.join(acct_parts),
            research_links=website if website.startswith('http') else '',
            contacts=contacts,
        ))

    return write_p2_csv('source_p2_government_congress.csv', rows)


def convert_nyc_initiatives(wb):
    """Convert NYC Initiatives tab.

    Structure: rows with a Type name are initiative headers; rows without a Type
    but with a Contact are members of the preceding initiative. We group contacts
    under their initiative and create one account per initiative. Rows that have
    a Type but no contacts (like standalone programs) become their own account.
    """
    ws = wb['Gov-NYC Initiatives ']

    # First pass: group rows by initiative
    initiatives = []  # list of {type, contacts: [(name, pos, email, addl)], status, outreach, notes}
    current = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(c for c in row if c is not None and str(c).strip()):
            continue
        init_type = cell(row[0])
        contact_name = cell(row[1])
        position = cell(row[2])
        email = first_email(cell(row[3]))
        addl_contact = clean_multiline(row[4])
        status = cell(row[5])
        outreach = clean_multiline(row[6])
        notes = clean_multiline(row[7])

        if init_type:
            # New initiative header
            current = {
                'type': init_type,
                'contacts': [],
                'status': status,
                'outreach': outreach,
                'notes': notes,
            }
            initiatives.append(current)
            if contact_name:
                current['contacts'].append((contact_name, position, email))
            if addl_contact:
                addl_email = first_email(addl_contact)
                addl_name = re.sub(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', '', addl_contact)
                addl_name = re.sub(r'[,;]\s*$', '', addl_name).strip()
                if addl_name and len(addl_name) > 2:
                    current['contacts'].append((addl_name, '', addl_email))
        elif contact_name and current:
            # Member row under the current initiative
            current['contacts'].append((contact_name, position, email))
        elif contact_name:
            # Orphan contact with no initiative header — create standalone
            initiatives.append({
                'type': f"NYC Contact - {contact_name}",
                'contacts': [(contact_name, position, email)],
                'status': status,
                'outreach': outreach,
                'notes': notes,
            })

    # Second pass: build Phase 2 rows
    rows = []
    for init in initiatives:
        name = f"NYC Gov - {init['type']}"
        ext_id = slugify(name)
        contacts = init['contacts'][:3]  # Max 3 contacts per account

        # If initiative has >3 contacts, capture extra names in notes
        extra_contacts = init['contacts'][3:]
        notes = init['notes']
        if extra_contacts:
            extra_names = ', '.join(c[0] for c in extra_contacts)
            notes = f"{notes}; Additional members: {extra_names}" if notes else f"Additional members: {extra_names}"

        rows.append(make_row(
            ext_id=ext_id,
            name=name,
            source_sheet='Government NYC Initiatives',
            location='New York City',
            app_status=init['status'],
            outreach=init['outreach'],
            notes=notes,
            contacts=contacts,
        ))

    return write_p2_csv('source_p2_government_nyc_initiatives.csv', rows)


def convert_borough_presidents(wb):
    """Convert Current Borough President tab.

    5 borough presidents with chief of staff contacts.
    """
    ws = wb['Gov-Current Borough President']
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(c for c in row if c is not None and str(c).strip()):
            continue
        borough = cell(row[0])
        president = cell(row[1])
        pres_email = first_email(cell(row[2]))
        phone = cell(row[3])
        cos = cell(row[4])
        cos_email = first_email(cell(row[5]))
        scheduling = clean_multiline(row[6])
        charter_pos = cell(row[7])
        funding_link = cell(row[8]) if len(row) > 8 else ''

        if not president:
            continue

        name = f"{borough} Borough President - {president}"
        ext_id = slugify(name)

        contacts = [(president, 'Borough President', pres_email)]
        if cos:
            contacts.append((cos, 'Chief of Staff', cos_email))

        acct_parts = []
        if phone:
            acct_parts.append(f"Phone: {phone}")
        if scheduling:
            acct_parts.append(f"Scheduling: {scheduling}")
        if charter_pos:
            acct_parts.append(f"Charter School Position: {charter_pos}")

        rows.append(make_row(
            ext_id=ext_id,
            name=name,
            source_sheet='Government Borough Presidents',
            location=f"{borough}, New York City",
            acct_info='; '.join(acct_parts),
            research_links=funding_link if funding_link and 'http' in str(funding_link).lower() else '',
            notes=f"Discretionary Funding: {funding_link}" if funding_link else '',
            contacts=contacts,
        ))

    return write_p2_csv('source_p2_government_borough_presidents.csv', rows)


# ── ERG Tab Converters ────────────────────────────────────────────────────────

def convert_erg_standard(wb, tab_name, source_suffix):
    """Convert a standard ERG tab (most ERG tabs share the same structure).

    Columns: Name, Area of interest, ERG Info, Contacts, Title, Email,
             Outreach, Status, Notes
    """
    ws = wb[tab_name]
    rows = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(c for c in row if c is not None and str(c).strip()):
            continue
        name = cell(row[0])
        area = cell(row[1])
        erg_info = clean_multiline(row[2])
        contacts_raw = cell(row[3])
        title_raw = cell(row[4])
        email_raw = cell(row[5])
        outreach = clean_multiline(row[6])
        status = cell(row[7])
        notes = clean_multiline(row[8]) if len(row) > 8 else ''

        if not name:
            continue

        ext_id = slugify(name)

        # Extract LinkedIn URL from email field if present
        email_clean, linkedin_url = extract_linkedin(email_raw)

        # Split multi-value contacts
        contacts = split_contacts(contacts_raw, title_raw, email_clean)

        # Build notes combining ERG info + status + notes
        notes_parts = []
        if erg_info:
            notes_parts.append(f"ERG: {erg_info}")
        if notes:
            notes_parts.append(notes)

        rows.append(make_row(
            ext_id=ext_id,
            name=name,
            source_sheet=f'ERG {source_suffix}',
            relationship='Employee Resource Group',
            area=area,
            app_status=status,
            outreach=outreach,
            notes='; '.join(notes_parts),
            research_links=linkedin_url,
            contacts=contacts,
        ))

    return rows


def convert_erg_partners(wb):
    """Convert ERG-Partners tab (has dual structure with ERG + Foundation columns).

    Uses only ERG columns: Name, Area, ERG Info, ERG Contacts, ERG Title, ERG Emails.
    """
    ws = wb['ERG-Partners']
    rows = []

    for row_data in ws.iter_rows(min_row=2, values_only=True):
        if not any(c for c in row_data if c is not None and str(c).strip()):
            continue
        name = cell(row_data[0])
        area = cell(row_data[1])
        erg_info = clean_multiline(row_data[2])
        erg_contacts = cell(row_data[3])
        erg_title = cell(row_data[4])
        erg_emails = cell(row_data[5])
        updates = clean_multiline(row_data[6])

        if not name:
            # Check if this is a section header row (like "BEAUTY AND PERSONAL CARE")
            if area and not erg_contacts:
                continue  # Skip section headers
            continue

        ext_id = slugify(name)

        # Split multi-value contacts
        contacts = split_contacts(erg_contacts, erg_title, erg_emails)

        rows.append(make_row(
            ext_id=ext_id,
            name=name,
            source_sheet='ERG Partners',
            relationship='Employee Resource Group',
            area=area,
            notes=f"ERG: {erg_info}" if erg_info else '',
            outreach=updates,
            contacts=contacts,
        ))

    return rows


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print(f'Loading {RAW}...')
    wb = openpyxl.load_workbook(RAW, data_only=True)

    print('\n=== Government Tabs ===')

    # 1. Council Members (merge TOTAL + Outreach)
    convert_council_members(wb)

    # 2. Congress Members
    convert_congress_members(wb)

    # 3. NYC Initiatives
    convert_nyc_initiatives(wb)

    # 4. Borough Presidents
    convert_borough_presidents(wb)

    # Skip: Gov-Borough Delegations (0 data rows)
    # Skip: Gov-State Level (incomplete/fragment data)
    print('  Skipped: Gov-Borough Delegations (0 data rows)')
    print('  Skipped: Gov-State Level (incomplete fragment data)')

    print('\n=== ERG Tabs ===')

    # ERG-Partners (special dual structure)
    partner_rows = convert_erg_partners(wb)
    write_p2_csv('source_p2_erg_partners.csv', partner_rows)

    # Standard ERG tabs
    erg_standard_tabs = [
        ('ERG-Beauty and Personal Care', 'Beauty and Personal Care'),
        ('ERG-Communications', 'Communications'),
        ('ERG-Finance', 'Finance'),
        ('ERG-Food', 'Food'),
        ('ERG-Hospitality', 'Hospitality'),
        ('ERG-Insurance', 'Insurance'),
        ('ERG-Major Retail', 'Major Retail'),
        ('ERG-Tech', 'Tech'),
        ('ERG-Transportation', 'Transportation'),
        ('ERG-Wellness', 'Wellness'),
        ('ERG-Reattempt', 'Reattempt'),
        ('ERG-Not Applicable', 'Not Applicable'),
    ]

    for tab_name, suffix in erg_standard_tabs:
        tab_rows = convert_erg_standard(wb, tab_name, suffix)
        fname = f"source_p2_erg_{suffix.lower().replace(' ', '_').replace('and_', '')}.csv"
        write_p2_csv(fname, tab_rows)

    # Skip: ERG-IGNORE-SPLIT (0 data rows)
    print('  Skipped: ERG-IGNORE-SPLIT (0 data rows)')

    print('\n=== Done ===')


if __name__ == '__main__':
    main()
